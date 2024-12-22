import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os

# Load the Excel sheet or create a new one if it doesn't exist
def load_excel(file_path):
    try:
        if os.path.exists(file_path):
            # If the file exists, load it
            book = openpyxl.load_workbook(file_path)
            sheet = book['Sheet1']
        else:
            # If the file doesn't exist, create a new workbook and add a sheet
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = 'Sheet1'
            # Create column headers in the new sheet
            sheet['A1'] = 'Roll No'
            sheet['B1'] = 'Email'
            sheet['C1'] = 'CI Attendance'
            sheet['D1'] = 'Python Attendance'
            sheet['E1'] = 'DM Attendance'
            # Save the new workbook
            book.save(file_path)
        return sheet, book
    except Exception as e:
        print(f"Error loading or creating Excel file: {e}")
        return None, None

# Save the Excel file
def save_excel(book, file_path):
    try:
        book.save(file_path)
        print("File saved successfully!")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

# Send email to students
def send_email(recipients, subject, body, sender_email, sender_password):
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)

            for recipient in recipients:
                message = MIMEMultipart()
                message['Subject'] = subject
                message.attach(MIMEText(body, 'plain'))
                server.sendmail(sender_email, recipient, message.as_string())
        print(f"Emails sent successfully to {len(recipients)} students.")
    except Exception as e:
        print(f"Error sending email: {e}")

# Check attendance and send notifications
def check_attendance(sheet, subject_code, student_ids, max_allowed_leaves, sender_email, sender_password, staff_mails):
    row_num = []
    no_of_days = []
    students_to_warn = []
    staff_warn_msg = ""
    subject_name = ""

    # Map subject codes to names and warning messages
    subject_map = {
        1: ("CI", "Warning! You can take only one more day leave for CI class."),
        2: ("Python", "Warning! You can take only one more day leave for Python class."),
        3: ("DM", "Warning! You can take only one more day leave for DM class.")
    }
    
    subject_name, warning_msg = subject_map.get(subject_code, ("", ""))
    
    if not subject_name:
        print("Invalid subject code!")
        return

    # Process each student
    for student_id in student_ids:
        for i in range(2, sheet.max_row + 1):
            student_roll_no = sheet.cell(row=i, column=1).value
            if student_roll_no == student_id:
                attendance_col = subject_code + 2  # column 3 for CI, 4 for Python, 5 for DM
                current_attendance = sheet.cell(row=i, column=attendance_col).value or 0
                new_attendance = current_attendance + 1
                sheet.cell(row=i, column=attendance_col).value = new_attendance
                no_of_days.append(new_attendance)
                row_num.append(i)

                # Check if attendance exceeds limit
                if new_attendance == max_allowed_leaves:
                    students_to_warn.append(sheet.cell(row=i, column=2).value)
                    send_email([sheet.cell(row=i, column=2).value], 'Attendance Warning', warning_msg, sender_email, sender_password)
                elif new_attendance > max_allowed_leaves:
                    staff_warn_msg += f"Roll No: {student_roll_no} - {sheet.cell(row=i, column=2).value}\n"

    # Send email to staff
    if staff_warn_msg:
        send_email(staff_mails, f"Lack of Attendance in {subject_name}", staff_warn_msg, sender_email, sender_password)

# Main function to handle user input and attendance updates
def main():
    file_path = 'D:\\laragon\\www\\python\\Learn\\Attendenc\\attendance.xlsx'
    sheet, book = load_excel(file_path)
    if not sheet:
        return

    sender_email = os.getenv("SENDER_EMAIL", "crazygirlaks@gmail.com")
    sender_password = os.getenv("SENDER_PASSWORD", "ERAkshaya485")  # Use environment variables for security
    staff_mails = ['muajjam.imu@gmail.com', 'test.muajjam@gmail.com']
    
    while True:
        print("1--->CI\n2--->Python\n3--->DM")
        subject_code = int(input("Enter subject: "))
        
        max_allowed_leaves = 2
        no_of_absentees = int(input('Number of absentees: '))

        student_ids = []
        if no_of_absentees > 1:
            student_ids = list(map(int, input('Enter roll numbers (separated by space): ').split()))
        else:
            student_ids = [int(input('Enter roll number: '))]

        # Update attendance and send warnings
        check_attendance(sheet, subject_code, student_ids, max_allowed_leaves, sender_email, sender_password, staff_mails)
        
        save_excel(book, file_path)

        resp = int(input('Another subject? 1--->yes 0--->no: '))
        if resp == 0:
            break

if __name__ == "__main__":
    main()
